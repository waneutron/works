from flask import Flask, render_template, request, redirect, url_for, jsonify
import sqlite3
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Configuration for file uploads
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create upload folder if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_db():
    """Initialize database with schema if it doesn't exist"""
    conn = sqlite3.connect('library.db')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS books (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            author TEXT NOT NULL,
            isbn TEXT UNIQUE NOT NULL,
            quantity INTEGER NOT NULL,
            room TEXT,
            shelf TEXT,
            section TEXT
        )
    ''')
    conn.commit()
    conn.close()

def get_db_connection():
    """Get database connection with error handling"""
    try:
        # Initialize database on first run
        init_db()
        conn = sqlite3.connect('library.db')
        conn.row_factory = sqlite3.Row
        # Test connection
        conn.execute('PRAGMA integrity_check')
        return conn
    except sqlite3.DatabaseError as e:
        # Database is corrupted, recreate it
        import os
        if os.path.exists('library.db'):
            os.rename('library.db', 'library.db.corrupted')
        init_db()
        conn = sqlite3.connect('library.db')
        conn.row_factory = sqlite3.Row
        return conn

@app.route('/')
def index():
    conn = get_db_connection()
    books = conn.execute('SELECT * FROM books').fetchall()
    conn.close()
    return render_template('index.html', books=books)

@app.route('/add', methods=('GET', 'POST'))
def add_book():
    if request.method == 'POST':
        title = request.form['title']
        author = request.form['author']
        isbn = request.form['isbn']
        quantity = request.form['quantity']
        room = request.form['room']
        shelf = request.form['shelf']
        section = request.form['section']
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO books (title, author, isbn, quantity, room, shelf, section) 
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (title, author, isbn, quantity, room, shelf, section))
        conn.commit()
        conn.close()
        return redirect(url_for('index'))
    return render_template('add.html')

@app.route('/edit/<int:id>', methods=('GET', 'POST'))
def edit_book(id):
    conn = get_db_connection()
    book = conn.execute('SELECT * FROM books WHERE id = ?', (id,)).fetchone()
    if request.method == 'POST':
        title = request.form['title']
        author = request.form['author']
        isbn = request.form['isbn']
        quantity = request.form['quantity']
        room = request.form['room']
        shelf = request.form['shelf']
        section = request.form['section']
        conn.execute('''
            UPDATE books SET title=?, author=?, isbn=?, quantity=?, room=?, shelf=?, section=? 
            WHERE id=?
        ''', (title, author, isbn, quantity, room, shelf, section, id))
        conn.commit()
        conn.close()
        return redirect(url_for('index'))
    conn.close()
    return render_template('edit.html', book=book)

@app.route('/delete/<int:id>', methods=['POST'])
def delete_book(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM books WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

# --- Route untuk carian segera (live search) ---
@app.route('/search')
def search():
    query = request.args.get('q', '')
    conn = get_db_connection()
    books = conn.execute("""
        SELECT * FROM books 
        WHERE title LIKE ? OR author LIKE ? OR isbn LIKE ?
    """, (f'%{query}%', f'%{query}%', f'%{query}%')).fetchall()
    conn.close()
    results = [dict(book) for book in books]
    return jsonify(results)

@app.route('/import', methods=('GET', 'POST'))
def import_excel():
    if request.method == 'POST':
        # Check if file is present
        if 'file' not in request.files:
            return redirect(url_for('index'))
        
        file = request.files['file']
        
        if file.filename == '':
            return redirect(url_for('index'))
        
        if not allowed_file(file.filename):
            return redirect(url_for('index'))
        
        try:
            # Save the file temporarily
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Load the Excel file
            wb = load_workbook(filepath)
            ws = wb.active
            
            conn = get_db_connection()
            inserted_count = 0
            
            # Skip header row and process data
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is None:  # Skip empty rows
                    continue
                
                try:
                    title = row[0] or ''
                    author = row[1] or ''
                    isbn = row[2] or ''
                    quantity = row[3] or 0
                    room = row[4] or ''
                    shelf = row[5] or ''
                    section = row[6] or ''
                    
                    conn.execute('''
                        INSERT INTO books (title, author, isbn, quantity, room, shelf, section) 
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (title, author, isbn, quantity, room, shelf, section))
                    inserted_count += 1
                except Exception as e:
                    # Log error and continue with next row
                    print(f"Error inserting row {row_idx}: {str(e)}")
                    continue
            
            conn.commit()
            conn.close()
            
            # Clean up - delete the uploaded file
            os.remove(filepath)
            
            return redirect(url_for('index'))
        except Exception as e:
            print(f"Error processing Excel file: {str(e)}")
            return redirect(url_for('index'))
    
    return render_template('import.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
